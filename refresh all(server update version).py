from dependencies.ServerHandler import ServerHandler
from dependencies.FileHandler import FileHandler
from dependencies.ETLHandler import ETLHandler
from dependencies.ServerHandler import CoreExportServerHandler
import pandas as pd
import concurrent.futures
import math
import sys
import pysftp
import paramiko
from openpyxl import load_workbook
import base64
from multiprocessing import Pool
import os
import datetime
import pdb
import time
config_path = 'dependencies/config.xlsx'
config_file = pd.read_excel(config_path,sheet_name='basic info')
dimension_query_result_folder_path = r'\\123\dimension query result'
blocker_info_query_result_folder_path = r'\\123\blocker info query result'

country_code_mapping_df = pd.read_excel(config_path, sheet_name='country code mapping')
country_code_mapping_dic = {}
code_country_mapping_dic = {}
for index, row in country_code_mapping_df.iterrows():
    country_code_mapping_dic.update({row['country']: row['country code']})
    code_country_mapping_dic.update({row['country code']: row['country']})


def wait_until_having_internet(source,asin_list,try_time,task):
    count=0
    while count < try_time:
        try:
            time.sleep(60)
            if task == 'refresh dimension':
                serverhandler = ServerHandler(database='notconveyable_'+source.lower())
                serverhandler.update_dimension_values(asin_list)
        except:
            time.sleep(60)
            serverhandler = []
        count+=1
    if serverhandler !=[]:
        return serverhandler
    else:
        print('已经在'+str(try_time*2)+'分钟内尝试'+str(try_time)+'次连接，均连接不上服务器，请检查网络')
def get_current_progress(finished_unit, all_unit, source):
    percentage = round((finished_unit / all_unit) * 100, 2)
    return source + '更新数据库进度：' + str(percentage) + '%'
def refresh_dimension(source_list_list,asin_list_list_by_source,query_list_list_by_source,chrome_profile_path,start_from_scratch_flag):
    '''
    跑完一个source的query，然后点完prioritize的结果后再去跑另一个source的query，直到所有source的query都跑完
    等待query结果，然后从共享中抓取后
    :param source_list_list: 对在list中的source进行操作；根据config中的basic info中的第一列获取；[['US','US'],['DE','DE','DE']]
    :param asin_list_list_by_source: 存放的是每个source要进行操作的ASIN list的list，每个最内层的ASIN list对应一个query link
    [[['ASIN1','ASIN2'],['ASIN11','ASIN12','ASIN13']],[['ASIN21','ASIN22'],['ASIN23'],['ASIN24']]]
    :param query_list_list_by_source: 存放的是每个source对应的query，query个数取决于source_list中内层source的个数
    :param chrome_profile_path: 一个user date文件夹地址
    :return: 返回一个待更新的数据，然后汇总到一个大的df中，根据最后的汇总df，对数据库中记录进行统一更新
    '''
    process_start_time = time.time()
    # source_list_list = [['US','US'],['DE','DE','DE']]
    # asin_list_list_by_source = [[['B002PU9WGQ','B071D994ND','B00J2M1YXE','B001B4VGFC'],['B0061WDZ9E','B06ZY3FNLQ']],[['B0098J4NMU','B002TUSRWW','B003AWMXZA'],['B01LD2BX36','B01N02W32X'],['B01964ZDHA','B0136RY72U']]]
    # query_list_list_by_source = [['https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256095','https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256098'],
    #                         ['https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256151','https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256162','https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256166']]
    # 如果从头运行程序的flag的True，则用国家，asin，querylist去查blocker status
    if start_from_scratch_flag == True:
        # source_list_list = [['US'], ['DE', 'DE', 'DE']]
        # asin_list_list_by_source = [
        #     [['B07B3T37G8', 'B07D5CMLF7','B07H7W9SSR','B07HDGQYTM','B07KYTQ1GZ','B07S1NRTXN','B07X1T8QGD','B07Z7BJYRH','B082J3PYBM','B083F43WQZ','B084VNSK7P','B084Z5CCHR','B085791MD9']],
        #     [['B00004YV4U', 'B00006J6O8', 'B00008JNUY'], ['B00007EEEK', 'B000092QSA','B0000AI3LP','B0000AITKK','B0002S4P1U','B0002YYNWA'], ['B00009QH6G', 'B00009WKHR','B000L9M508','B000LXXPSA','B000ONNOLK','B000UW2XPI','B000VQR6LY','B000WL4UR6','B000WOSMOA','B0014DT666']]]
        # query_list_list_by_source = [
        #     [
        #      'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256098'],
        #     ['https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256151',
        #      'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256162',
        #      'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10256166']]
        query_job_list_list_by_source = []
        good_query_status_list_list_by_source = []
        is_query_err_dic = {}
        is_query_complete_list_list = []
        query_type = 'refresh dimension'
        #每个source的query排队跑一遍，点prioritize
        for index1,source_list in enumerate(source_list_list):
            query_job_list_by_source = []
            good_query_status_list_by_source = []
            is_query_complete_list = []
            for index2, source in enumerate(source_list):
                # pdb.set_trace()
                etlhandler = ETLHandler(chrome_profile_path)
                print(source)
                print(asin_list_list_by_source[index1][index2])
                print(query_list_list_by_source[index1][index2])
                print(query_type)
                link, good_query_status = etlhandler.auto_run_sql(source,asin_list_list_by_source[index1][index2],query_list_list_by_source[index1][index2],query_type)
                query_job_list_by_source.append(link)
                good_query_status_list_by_source.append(good_query_status)
                is_query_complete_list.append(False)
                print(query_job_list_by_source)
                print(good_query_status_list_by_source)
            query_job_list_list_by_source.append(query_job_list_by_source)
            good_query_status_list_list_by_source.append(good_query_status_list_by_source)
            is_query_complete_list_list.append(is_query_complete_list)
            if False not in good_query_status_list_by_source:
                is_query_err_dic.update({source_list[0]:False})
            else:
                is_query_err_dic.update({source_list[0]: True})
        # 将is_query_complete_list_list和is_query_err_dic和query_job_list_list_by_source记录到一个excel里，留作程序报错时直接将已经进行的query交给程序的入口。
        history_df = pd.DataFrame(data={'is_query_complete_list_list': [str(is_query_complete_list_list)],
                                        'is_query_err_dic': [str(is_query_err_dic)],
                                        'query_job_list_list_by_source': [str(query_job_list_list_by_source)]})

        book = load_workbook(config_path)
        writer = pd.ExcelWriter(config_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        history_df.to_excel(writer, sheet_name='dimension query history',index=False)
        writer.save()
    # 如果从头运行程序的flag为false，则读取history_df中的is_query_complete_list_list；is_query_err_dic；query_job_list_list_by_source信息
    else:


        history_df = pd.read_excel(config_path, sheet_name='dimension query history')
        is_query_complete_list_list = eval(history_df.loc[0, 'is_query_complete_list_list'])
        is_query_err_dic = eval(history_df.loc[0, 'is_query_err_dic'])
        query_job_list_list_by_source = eval(history_df.loc[0, 'query_job_list_list_by_source'])
        # etlhandler = ETLHandler(chrome_profile_path)
    # print(query_job_list_list_by_source)
    #[['https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285199752'], ['https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285204844', 'https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285213759', 'https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285224514']]
    # print(good_query_status_list_list_by_source)
    #[[True], [True, True, True]]
    # print(is_query_err_dic)
    #{'US': False, 'DE': False}
    #查看query的结果，每10分钟看一个query link是否结束，若结束则去看下一个query link直到全部query link都看完
    print('is_query_complete_list_list的值为',is_query_complete_list_list)
    # print(source_list_list)
    for index1, source_list in enumerate(source_list_list):
        index2 = 0
        while False in is_query_complete_list_list[index1]:
            # 如果query中一开始有报错的情况，则将is_query_complete_list_list[index1]中的所有元素全更新成True
            if is_query_err_dic[source_list[0]] == True:
                for i in range(len(is_query_complete_list_list[index1])):
                    is_query_complete_list_list[index1][i] = True

                continue
            # 如果query一开始没有报错，则每隔5分组去检查query有没有成功
            else:
                etlhandler = ETLHandler(chrome_profile_path)
                good_query_status,is_query_finish = etlhandler.wait_query_and_get_result_after(query_link=query_job_list_list_by_source[index1][index2])
                #如果任意一个query报错，则将is_query_complete_list_list[index1]所有的元素标记为True，不再检查此source的query
                if good_query_status == False:
                    is_query_err_dic[source_list[0]] = True
                    for i in range(len(is_query_complete_list_list[index1])):
                        is_query_complete_list_list[index1][i] = True
                    index2 += 1
                #如果当前query没有报错，则看当前query是否结束，若结束则将is_query_complete_list_list[index1][index2]标记为True
                else:
                    if is_query_finish == True:
                        is_query_complete_list_list[index1][index2] = True
                        index2 += 1
                    else:
                        time.sleep(300)
    #如果其中一个国家的结果有出来，则对其进行后面的操作，如果没有任何一个国家有结果，则不进行后面的拿结果的操作
    if False in is_query_err_dic.values():
        #拿到所有没报错的query结果地址
        all_available_file_path_by_source_dic = {}
        today = datetime.timedelta(seconds=time.time()).days
        for k,v in is_query_err_dic.items():
            if v == False:
                all_available_file_path_by_source_dic.update({k:[]})
                full_dimension_query_result_folder_path = '/'.join([dimension_query_result_folder_path,k])
                file_name_list = os.listdir(full_dimension_query_result_folder_path)
                for file_name in file_name_list:
                    full_name_path = '/'.join([full_dimension_query_result_folder_path,file_name])
                    if datetime.timedelta(seconds=os.path.getmtime(full_name_path)).days ==today:
                        all_available_file_path_by_source_dic[k].append(full_name_path)
            # pdb.set_trace()
        #读取结果
        all_source_df = pd.DataFrame()
        for k,v in all_available_file_path_by_source_dic.items():
            for path in v:
                temp_df = pd.read_csv(path,error_bad_lines=False,sep='\t')
                all_source_df = all_source_df.append(temp_df)
            # pdb.set_trace()

        # all_source_df.to_excel('final_result_to_update_db.xlsx',index=False)
        #将结果上传到数据库上，等待更新
        all_source_df.reset_index(drop=True,inplace=True)
        all_source_df.fillna('',inplace=True)
        all_source_df['item_package_weight'] = ''
        all_source_df['item_package_dimensions'] = ''
        for index,row in all_source_df.iterrows():
            if row['pkg_height'] !='' and row['pkg_length']!='' and row['pkg_width']!='':
                all_source_df.loc[index,'item_package_dimensions'] = 'True'
            else:
                all_source_df.loc[index, 'item_package_dimensions'] = 'False'
            if row['pkg_weight'] !='':
                all_source_df.loc[index, 'item_package_weight'] = 'True'
            else:
                all_source_df.loc[index, 'item_package_weight'] = 'False'
        all_source_df.rename(columns={'marketplace_id':'source'},inplace=True)
        all_source_df['source'] = all_source_df['source'].map(lambda x:code_country_mapping_dic[x])
        all_source_df.to_excel('final_dimension_result_to_update_db.xlsx',index=False)
        print('保存dimension结果成功')
        # all_source_list = all_source_df['source'].unique().tolist()
        result_path = '/'.join([os.getcwd(), 'final_dimension_result_to_update_db.xlsx'])
        host = '123'
        # 进入服务器后台 ssh metron@123com
        user_name = 'metron'
        password = base64.b64decode('123').decode("utf-8", "ignore")
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        sftp_conn = pysftp.Connection(host, username=user_name, cnopts=cnopts, private_key='id_rsa')
        # sftp_conn.chdir('')
        sftp_conn.put(localpath=result_path,
                      remotepath='/home/metron/xinyyang/dimension_update/final_dimension_result_to_update_db.xlsx')
        sftp_conn.close()
        print('结果已经上传到服务器上')
        # pdb.set_trace()
        #@todo under construction

    #     prompt = ''
    #     print('更新数据库中......')
    #
    #     for source in all_source_list:
    #         #把待更新的数据改成1000个一组，然后多组进行更新，为了速度快一些
    #         # temp_df_by_source = all_source_df[all_source_df['source']==source][
    #         #     ['item_package_dimensions', 'item_package_weight', 'asin', 'source']]
    #         temp_df_by_source = all_source_df[all_source_df['source'] == source][
    #             ['item_package_dimensions', 'item_package_weight', 'asin', 'source']]
    #         temp_df_by_source.reset_index(drop=True,inplace=True)
    #         # asin_dimensional_info_list = all_source_df[all_source_df['source']==source][
    #         #     ['item_package_dimensions', 'item_package_weight', 'asin', 'source']].to_numpy().tolist()
    #         asins_by_source_num = temp_df_by_source.shape[0]
    #         unit_quality = 1000
    #         unit_num = math.ceil( asins_by_source_num/unit_quality)
    #         temp_df_list = []
    #         start_time = time.time()
    #         for index in range(unit_num):
    #             temp_df_list.append(temp_df_by_source.loc[index*unit_quality:(index+1)*unit_quality])
    #         # print(temp_df_list)
    #         serverhandler = ServerHandler(database='notconveyable_'+source.lower())
    #         for index in range(len(temp_df_list)):
    #             asin_dimensional_info_list = temp_df_list[index].to_numpy().tolist()
    #             print(get_current_progress(index,len(temp_df_list),source))
    #             #@todo 设置网络断开后重新尝试建立连接并且传数据到服务器上
    #             try:
    #                 serverhandler.update_dimension_values(asin_dimensional_info_list)
    #             except:
    #                 serverhandler.close_connection()
    #                 serverhandler = wait_until_having_internet(source.lower(),asin_dimensional_info_list,try_time=5,task='refresh dimension')
    #
    #         serverhandler.close_connection()
    #         end_time = time.time()
    #         print(end_time-start_time)
    #         prompt += (str(source) + '更新dimension相关属性成功，共有' + str(asins_by_source_num) + '个asin被更新了dimension相关属性;\n')
    #     print('更新数据库完毕......')
    #     if prompt !='':
    #         print(prompt)
    #     else:
    #         print('refresh dimension失败，请重试或联系xinyyang')
    # process_end_time = time.time()
    # print('process executed for '+str(process_end_time-process_start_time)+' seconds')
    # @todo under construction

def refresh_exportable_date(source_list):
    #一个source一个source的进行：找到代查询ASIN，然后去coreexport数据库中查询，然后更新到database1中
    # pdb.set_trace()
    print(source_list)
    refresh_country_list = pd.read_excel(config_path,sheet_name='refresh exportable date')['refresh country'].tolist()
    # country_code_mapping_df = pd.read_excel(config_path,sheet_name='country code mapping')
    # country_code_mapping_dic = {}
    # code_country_mapping_dic = {}
    # for index,row in country_code_mapping_df.iterrows():
    #     country_code_mapping_dic.update({row['country']:row['country code']})
    #     code_country_mapping_dic.update({row['country code']:row['country']})
    # pdb.set_trace()
    df_to_update_exportable_date_all_source = pd.DataFrame()
    is_updated_dic = {source:False for source in source_list}
    asins_with_exportable_date_df = pd.DataFrame()
    for source in source_list:
        if source in refresh_country_list:
            #找到待查ASIN
            # print(source)
            serverhandler = ServerHandler(database='notconveyable_'+source.lower())
            df_by_source = serverhandler.get_data_to_refresh()
            # pdb.set_trace()
            df_by_source['source'] = df_by_source['source'].map(lambda x: str(country_code_mapping_dic[x]))
            df_by_source['asinsource'] = df_by_source['asin']+df_by_source['source']
            asin_source_list = df_by_source['asinsource'].unique().tolist()
            print(asin_source_list)
            #去coreexport数据库中查询exportable_date
            #会出现数据库中抓不到记录导致报错的情况where concat(asin,marketplace_id) in ()中没有记录
            try:
                df_from_coreexport_by_source = CoreExportServerHandler(database='coreexport_'+source.lower()).get_exportable_data(asin_source_list)
            except:
                continue
            # pdb.set_trace()
            # print(df_from_coreexport_by_source)
            asins_with_exportable_date_df = asins_with_exportable_date_df.append(df_from_coreexport_by_source,ignore_index=True)
            df_from_coreexport_by_source.rename(columns={'marketplace_id':'source'},inplace=True)
            df_from_coreexport_by_source['source'] = df_from_coreexport_by_source['source'].map(lambda x: str(x))
            # df_from_coreexport_by_source['source'] = df_from_coreexport_by_source['source'].map(lambda x:country_code_mapping_dic[x])
            # pdb.set_trace()
            df_to_update_exportable_date = df_by_source.merge(df_from_coreexport_by_source,how='inner',on=['asin','source'])
            # pdb.set_trace()
            #更新not conveyable数据库
            df_to_update_exportable_date['source'] = df_to_update_exportable_date['source'].map(lambda x: int(x))
            df_to_update_exportable_date['source'] = df_to_update_exportable_date['source'].map(lambda x: code_country_mapping_dic[x])
            df_to_update_exportable_date = df_to_update_exportable_date[['exportable_date','asin','source']]
            df_to_update_exportable_date['exportable_date'] = df_to_update_exportable_date['exportable_date'].map(lambda x:x.strftime('%Y-%m-%d'))
            # asin_with_exportable_info_list = df_to_update_exportable_date.to_numpy().tolist()
            #改成分批量更新的那种情况
            asins_by_source_num = df_to_update_exportable_date.shape[0]
            unit_quality = 1000
            unit_num = math.ceil(asins_by_source_num / unit_quality)
            temp_df_list = []
            start_time = time.time()
            for index in range(unit_num):
                temp_df_list.append(df_to_update_exportable_date.loc[index * unit_quality:(index + 1) * unit_quality])
            # print(temp_df_list)
            # serverhandler = ServerHandler(database='notconveyable_' + source.lower())
            for index in range(len(temp_df_list)):
                asin_with_exportable_info_list = temp_df_list[index].to_numpy().tolist()
                print(get_current_progress(index, len(temp_df_list), source))
                serverhandler.update_data_to_refresh(asin_with_exportable_info_list)

            serverhandler.close_connection()
            # serverhandler.update_data_to_refresh(asin_with_exportable_info_list)
            is_updated_dic[source] = True
            df_to_update_exportable_date_all_source = df_to_update_exportable_date_all_source.append(df_to_update_exportable_date,ignore_index=True)
        else:
            print('此source:'+source+'不更新exportable date')
    # for index,source in enumerate(source_list):
    #更新完了以后输出哪些source更新了
    df_to_update_exportable_date_all_source.to_excel('asins update exportable date.xlsx',index=False)
    asins_with_exportable_date_df.to_excel('asins with exportable date.xlsx',index=False)
    prompt = ''
    asin_num_by_source_dic = get_operation_num(df_to_update_exportable_date_all_source,source_list)
    print(is_updated_dic)
    print(asin_num_by_source_dic)
    for k,v in is_updated_dic.items():
        if v == True:
            # print(k)
            prompt += (str(k)+'更新exportable date成功，共有'+str(asin_num_by_source_dic[k])+'个asin被更新了exportable date;\n')
        else:
            prompt += (str(k) + '更新exportable date失败，或者数据库中没有符合条件且需要更新的asin，请检查，或重新更新;\n')
    if prompt !='':
        print(prompt)
    else:
        print('没有source更新exportable date成功，请重试或联系xinyyang')
    # print(prompt)
# refresh_exportable_date(['US','DE','JP','UK'])
#@todo 用大批量的ASIN测试一下结果，小批量ASIN测试的结果目前没问题
def refresh_blocker_status(source_list_list,asin_list_list_by_source,query_list_list_by_source,chrome_profile_path,start_from_scratch_flag):

    #如果从头运行程序的flag的True，则用国家，asin，querylist去查blocker status
    if start_from_scratch_flag == True:
        # source_list_list = [['US'], ['DE', 'DE', 'DE']]
        # asin_list_list_by_source = [
        #     [['B09PJYXW84', 'B09P3JHJ78','B06XWHQSP3','B09L3B778Y','B09PGDQTQ3','B09P1C6XZQ','B08595QWB7','B09JYVJC4W']],
        #     [['B08NQ1FNW8', 'B09PFWZYM4', 'B09MXHY29B'], ['B09PB44P1G', 'B09LV4G1BR'], ['B084Z1THNH', 'B09B26VVTF']]]
        # query_list_list_by_source = [
        #     [
        #         'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10271809'],
        #     ['https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10271832',
        #      'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10271833',
        #      'https://datacentral.a2z.com/dw-platform/servlet/dwp/template/EtlViewExtractJobs.vm/job_profile_id/10271834']]
        query_job_list_list_by_source = []
        # good_query_status_list_list_by_source = []
        is_query_err_dic = {}
        is_query_complete_list_list = []
        query_type = 'refresh blocker status'
        # 每个source的query排队跑一遍，点prioritize
        print("user-data-dir=" +chrome_profile_path)
        for index1, source_list in enumerate(source_list_list):
            query_job_list_by_source = []
            good_query_status_list_by_source = []
            is_query_complete_list = []
            for index2, source in enumerate(source_list):
                # pdb.set_trace()
                etlhandler = ETLHandler(chrome_profile_path)
                link, good_query_status = etlhandler.auto_run_sql(source, asin_list_list_by_source[index1][index2],
                                                                  query_list_list_by_source[index1][index2],query_type)
                query_job_list_by_source.append(link)
                good_query_status_list_by_source.append(good_query_status)
                is_query_complete_list.append(False)
                print(query_job_list_by_source)
                print(good_query_status_list_by_source)
            query_job_list_list_by_source.append(query_job_list_by_source)
            # good_query_status_list_list_by_source.append(good_query_status_list_by_source)
            is_query_complete_list_list.append(is_query_complete_list)
            if False not in good_query_status_list_by_source:
                is_query_err_dic.update({source_list[0]: False})
            else:
                is_query_err_dic.update({source_list[0]: True})
        #将is_query_complete_list_list和is_query_err_dic和query_job_list_list_by_source记录到一个excel里，留作程序报错时直接将已经进行的query交给程序的入口。
        history_df = pd.DataFrame(data={'is_query_complete_list_list':[str(is_query_complete_list_list)],'is_query_err_dic':[str(is_query_err_dic)],'query_job_list_list_by_source':[str(query_job_list_list_by_source)]})
        #把新的更改存一下
        # history_df.to_excel(config_path,sheet_name='blocker query history')
        book = load_workbook(config_path)
        writer = pd.ExcelWriter(config_path, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        history_df.to_excel(writer, sheet_name='blocker query history',index=False)
        writer.save()
    #如果从头运行程序的flag为false，则读取history_df中的is_query_complete_list_list；is_query_err_dic；query_job_list_list_by_source信息
    else:
        history_df = pd.read_excel(config_path,sheet_name='blocker query history')
        is_query_complete_list_list = eval(history_df.loc[0,'is_query_complete_list_list'])
        is_query_err_dic = eval(history_df.loc[0,'is_query_err_dic'])
        query_job_list_list_by_source = eval(history_df.loc[0,'query_job_list_list_by_source'])
    print(query_job_list_list_by_source)
    # [['https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285199752'], ['https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285204844', 'https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285213759', 'https://datacentral.a2z.com/console?action=jobrun_details&jobrun_id=7285224514']]
    # print(good_query_status_list_list_by_source)
    # [[True], [True, True, True]]
    print(is_query_err_dic)
    # {'US': False, 'DE': False}
    # 查看query的结果，每10分钟看一个query link是否结束，若结束则去看下一个query link直到全部query link都看完
    # is_query_err_dic['US'] = True
    print('is_query_complete_list_list的值为', is_query_complete_list_list)
    for index1, source_list in enumerate(source_list_list):
        index2 = 0
        while False in is_query_complete_list_list[index1]:
            # 如果query中一开始有报错的情况，则将is_query_complete_list_list[index1]中的所有元素全更新成True
            if is_query_err_dic[source_list[0]] == True:
                for i in range(len(is_query_complete_list_list[index1])):
                    is_query_complete_list_list[index1][i] = True

                continue
            # 如果query一开始没有报错，则每隔5分组去检查query有没有成功
            else:
                etlhandler = ETLHandler(chrome_profile_path)
                good_query_status, is_query_finish = etlhandler.wait_query_and_get_result_after(
                    query_link=query_job_list_list_by_source[index1][index2])
                # 如果任意一个query报错，则将is_query_complete_list_list[index1]所有的元素标记为True，不再检查此source的query
                if good_query_status == False:
                    is_query_err_dic[source_list[0]] = True
                    for i in range(len(is_query_complete_list_list[index1])):
                        is_query_complete_list_list[index1][i] = True
                    index2 += 1
                # 如果当前query没有报错，则看当前query是否结束，若结束则将is_query_complete_list_list[index1][index2]标记为True
                else:
                    if is_query_finish == True:
                        is_query_complete_list_list[index1][index2] = True
                        index2 += 1
                    else:
                        time.sleep(300)
    # 如果其中一个国家的结果有出来，则对其进行后面的操作，如果没有任何一个国家有结果，则不进行后面的拿结果的操作
    if False in is_query_err_dic.values():
        # 拿到所有没报错的query结果地址
        all_available_file_path_by_source_dic = {}
        today = datetime.timedelta(seconds=time.time()).days
        for k, v in is_query_err_dic.items():
            if v == False:
                all_available_file_path_by_source_dic.update({k: []})
                full_blocker_status_query_result_folder_path = '/'.join([blocker_info_query_result_folder_path, k])
                file_name_list = os.listdir(full_blocker_status_query_result_folder_path)
                for file_name in file_name_list:
                    full_name_path = '/'.join([full_blocker_status_query_result_folder_path, file_name])
                    #只拿当天的结果
                    if datetime.timedelta(seconds=os.path.getmtime(full_name_path)).days == today:
                        all_available_file_path_by_source_dic[k].append(full_name_path)
        #         pdb.set_trace()
        # pdb.set_trace()
        # 从共享中读取结果
        all_source_df = pd.DataFrame()
        for k, v in all_available_file_path_by_source_dic.items():
            # pdb.set_trace()
            for path in v:
                print(path)
                temp_df = pd.read_csv(path, error_bad_lines=False, sep='\t')
                all_source_df = all_source_df.append(temp_df)


        # all_source_df.to_excel('final_result_to_update_db_blocker.xlsx', index=False)
        # for key in all_available_file_path_by_source_dic.keys():
        #     serverhander = ServerHandler(database='notconveyable_'+key.lower())
        #     serverhander.get_data_to_delete_and_insert_to_history_table()

        #更新数据库
        blocker_mapping_df = pd.read_excel(config_path,'blocker mapping')
        # country_code_mapping_df = pd.read_excel(config_path, sheet_name='country code mapping')
        # country_code_mapping_dic = {}
        # code_country_mapping_dic = {}
        # for index, row in country_code_mapping_df.iterrows():
        #     country_code_mapping_dic.update({row['country']: row['country code']})
        #     code_country_mapping_dic.update({row['country code']: row['country']})
        #找到可以对数据库进行更改的所有source
        # available_source_list = all_source_df['marketplace_id'].map(lambda x:code_country_mapping_dic[x]).unique().tolist()

        #all_source_df为query的结果，下面将query的结果整理成特定的格式
        all_source_df.fillna(value='',inplace=True)
        all_source_df['blocker'] = ''
        # all_source_df['status'] = ''
        all_source_df.reset_index(drop=True,inplace=True)
        blocker_process_start_time = time.time()
        #在结果里面多加一列blocker列，blocker可能有Others; BizRule_ASINLevel_Too_Heavy; BizRule_ASINLevel_Too_Large和No blocker
        for index,row in all_source_df.iterrows():
            marketplace_id = row['marketplace_id']
            #'US'
            this_source = code_country_mapping_dic[marketplace_id]
            # pdb.set_trace()
            #如果reason里面有值，判断值里面是否含有config表中keyword列中的关键字，若有，则将blocker填成对应blocker_type中的值，若没有，则填成Others
            if row['contributions'] == 'OperationalRule' and row['restriction_reasons'] != '':

                blocker_mapping_df_by_source = blocker_mapping_df[blocker_mapping_df['source'] == this_source]
                total_count = blocker_mapping_df_by_source.shape[0]
                count = 0
                for index1, row1 in blocker_mapping_df_by_source.iterrows():

                    key_word_list = row1['keyword'].split(';')
                    if any(key_word in row['restriction_reasons'] for key_word in key_word_list):
                        all_source_df.loc[index,'blocker'] = row1['blocker_type']
                        break
                    count += 1

                if count == total_count:
                    all_source_df.loc[index, 'blocker'] = 'Others'

            elif row['contributions'] == 'OperationalRule' and row['restriction_reasons'] == '':
                all_source_df.loc[index, 'blocker'] = 'No blocker'

            else:
                all_source_df.loc[index, 'blocker'] = 'Others'
        # pdb.set_trace()
        all_source_df['asin'] = all_source_df['asin'].map(lambda x:str(x).zfill(10))
        all_source_df['marketplace_id'] = all_source_df['marketplace_id'].map(lambda x:code_country_mapping_dic[x])
        all_source_df.rename(columns={'marketplace_id':'source'},inplace=True)
        # print(all_source_df['source'])
        # pdb.set_trace()
        # pdb.set_trace()
        all_source_df['asinsource'] = all_source_df['asin'] + all_source_df['source']
        available_source_list = all_source_df['source'].unique().tolist()
        status_df = pd.DataFrame()
        #从数据库里把所有的ASIN的blocker和status信息抓出来
        for source in available_source_list:
            asin_source_list = all_source_df[all_source_df['source']==source]['asinsource'].tolist()
            serverhandler = ServerHandler(database='notconveyable_'+source.lower())
            # pdb.set_trace()
            status_df = status_df.append(serverhandler.get_status_info(asin_source_list),ignore_index=True)
            # pdb.set_trace()
        #all_source_df中的列有asin,source,blocker(从query结果中新查到的),blocker_name(db中来的),blocker_status
        all_source_df = all_source_df.merge(status_df,how='left',on=['asin','source'])
        # pdb.set_trace()
        all_source_df.drop_duplicates(keep='first',inplace=True,ignore_index=True,subset=['asin','source','blocker','restriction_reasons'])

        all_source_df['new_status'] = ''
        all_source_df['new_blocker'] = ''
        # all_source_df.to_excel('test blocker info.xlsx',index=False)
        for index,row in all_source_df.iterrows():

            #考虑一下一个asin有多个blocker name的情况,参考
            asin_count = all_source_df[all_source_df['asin']==row['asin']].shape[0]
            #用query查出来的ASIN只有单一blocker的情况
            if asin_count ==1:

                #数据库中原始blocker为A，新查出来blocker为B（nonactionable，如非toolarge tooheavy）
                if row['blocker'] == 'Others':
                    all_source_df.loc[index,'new_blocker'] = 'Others'
                    all_source_df.loc[index,'new_status'] = 'nonaction'
                    continue
                #如果数据库中原始blocker为others，但是新blocker为非others
                if row['blocker'] != 'Others' and row['blocker_name'] == 'Others':
                    all_source_df.loc[index, 'new_blocker'] = row['blocker']
                    all_source_df.loc[index, 'new_status'] = 'WIP'
                #数据库中原始blocker为A，新查出来的blocker为B（actionable，如toolarge或tooheavy）
                if row['blocker'] != 'Others' and row['blocker'] != row['blocker_name'] and row['blocker'] != 'No blocker':
                    all_source_df.loc[index, 'new_blocker'] = row['blocker']
                    all_source_df.loc[index, 'new_status'] = 'WIP'
                    continue
                #数据库中原始blocker为A，新查出来blocker消失，将新的new_blocker更新成数据库中的blocker name
                if row['blocker'] == 'No blocker':
                    all_source_df.loc[index, 'new_blocker'] = row['blocker_name']
                    all_source_df.loc[index, 'new_status'] = 'unblocked'
                    continue
                #数据库中原始blocker为A，新查出来blocker为A
                elif row['blocker'] == row['blocker_name']:
                    all_source_df.loc[index, 'new_blocker'] = row['blocker_name']
                    all_source_df.loc[index, 'new_status'] = 'WIP'
                    continue
            #若数据库中原始blocker为A，新blocker为A+B的情况
            elif asin_count >1:
                # pdb.set_trace()
                #有可能会出现相同asin但是有多个source的情况
                temp_df = all_source_df[all_source_df['asin'] == row['asin']]
                #对比一个asin是否有多个source
                asin_temp_source_list = temp_df['source'].unique().tolist()
                #如果多条相同ASIN的记录只有一个source
                if len(asin_temp_source_list) == 1:
                    temp_blocker_list = []
                    [temp_blocker_list.append(str(blocker_name)) for blocker_name in temp_df['blocker'].unique().tolist() if blocker_name not in temp_blocker_list]
                    #如果A是actionable的blocker，只要出现nonactionable的blocker，则把new_blocker更新成Others，status更新成nonaction
                    if 'Others' in temp_blocker_list:
                        all_source_df.loc[index, 'new_blocker'] = 'Others'
                        all_source_df.loc[index, 'new_status'] = 'nonaction'
                        # 有新查出来的blocker含有No blocker
                    elif 'No blocker' in temp_blocker_list:
                    # No blocker+现有blocker
                    # No blocker+新blocker
                        temp_blocker_list.remove('No blocker')
                        if temp_blocker_list !=[]:
                            all_source_df.loc[index, 'new_blocker'] = ','.join(temp_blocker_list)
                            all_source_df.loc[index, 'new_status'] = 'WIP'
                        else:
                            all_source_df.loc[index, 'new_blocker'] = row['blocker_name']
                            all_source_df.loc[index, 'new_status'] = 'unblocked'
                    #如果A是actionable的blocker且新查出来的A+B也都是actionble的blocker，则把new blocker更新成A+B
                    else:
                        all_source_df.loc[index, 'new_blocker'] = ','.join(temp_blocker_list)
                        all_source_df.loc[index, 'new_status'] = 'WIP'

                #如果多条ASIN的记录有多个source
                elif len(asin_temp_source_list) >1:
                    temp_source_list = temp_df['source'].unique().tolist()
                    #一个一个去看每个source的blocker情况
                    for temp_source in temp_source_list:
                        temp_df_by_source = temp_df[temp_df['source']==temp_source]

                        asin_count_by_source = temp_df_by_source.shape[0]
                        #每个source只有一个ASIN
                        if asin_count_by_source == 1:
                            # 数据库中原始blocker为A，新查出来blocker为B（nonactionable，如非toolarge tooheavy）
                            print(temp_df_by_source['blocker'])
                            if temp_df_by_source.loc[temp_df_by_source.index[0],'blocker'] == 'Others':
                                all_source_df.loc[index, 'new_blocker'] = 'Others'
                                all_source_df.loc[index, 'new_status'] = 'nonaction'
                                continue
                            # 数据库中原始blocker为A，新查出来blocker消失
                            if temp_df_by_source.loc[temp_df_by_source.index[0], 'blocker'] == 'No blocker':
                                all_source_df.loc[index, 'new_blocker'] = row['blocker_name']
                                all_source_df.loc[index, 'new_status'] = 'unblocked'
                                continue
                            # 数据库中原始blocker为A，新查出来的blocker为B（actionable，如toolarge或tooheavy）
                            if temp_df_by_source.loc[temp_df_by_source.index[0],'blocker'] != 'Others' and temp_df_by_source.loc[temp_df_by_source.index[0],'blocker'] != temp_df_by_source.loc[temp_df_by_source.index[0],'blocker_name']:
                                if temp_df_by_source.loc[temp_df_by_source.index[0],'blocker'] != 'No blocker':
                                    all_source_df.loc[index, 'new_blocker'] = temp_df_by_source.loc[temp_df_by_source.index[0],'blocker']
                                    all_source_df.loc[index, 'new_status'] = 'WIP'
                                else:
                                    all_source_df.loc[index, 'new_blocker'] = row['blocker_name']
                                    all_source_df.loc[index, 'new_status'] = 'unblocked'
                                continue

                            # 数据库中原始blocker为A，新查出来blocker为A
                            if temp_df_by_source.loc[temp_df_by_source.index[0],'blocker'] == temp_df_by_source.loc[temp_df_by_source.index[0],'blocker_name']:
                                all_source_df.loc[index, 'new_blocker'] = temp_df_by_source.loc[temp_df_by_source.index[0],'blocker_name']
                                all_source_df.loc[index, 'new_status'] = 'WIP'
                                continue
                        #如果一个asin有多条记录，也就是说有两个不同的blocker
                        elif asin_count_by_source >1:
                            #如果一个source的这个ASIN既有blocker A又有blocker B
                            temp_blocker_list = []
                            [temp_blocker_list.append(blocker_name) for blocker_name in
                             temp_df_by_source['blocker'].unique().tolist() if blocker_name not in temp_blocker_list]
                            # 如果A是actionable的blocker，只要出现nonactionable的blocker，则把new_blocker更新成Others，status更新成nonaction
                            if 'Others' in temp_blocker_list:
                                all_source_df.loc[index, 'new_blocker'] = 'Others'
                                all_source_df.loc[index, 'new_status'] = 'nonaction'
                            # 如果A是actionable的blocker且新查出来的A+B也都是actionble的blocker，则把new blocker更新成A+B
                            else:
                                all_source_df.loc[index, 'new_blocker'] = ','.join(temp_blocker_list)
                                all_source_df.loc[index, 'new_status'] = 'WIP'
            # pdb.set_trace()
        blocker_process_end_time = time.time()
        print('对比blocker用时: '+str(blocker_process_end_time-blocker_process_start_time))
        all_source_df.to_excel('final_result_to_update_db_new_blocker.xlsx', index=False)
        # pdb.set_trace()
        #@将结发布到server上
        result_path = '/'.join([os.getcwd(),'final_result_to_update_db_new_blocker.xlsx'])
        host = 'qwe'
        #进入服务器后台 ssh metron@qwe.us-west-2.amazon.com
        user_name = 'metron'
        password = base64.b64decode('qew').decode("utf-8", "ignore")
        cnopts = pysftp.CnOpts()
        cnopts.hostkeys = None
        sftp_conn = pysftp.Connection(host, username=user_name, cnopts=cnopts,private_key='id_rsa')
        # sftp_conn.chdir('')
        sftp_conn.put(localpath=result_path, remotepath='/home/metron/xinyyang/blocker_status_update/final_result_to_update_db_new_blocker.xlsx')
        sftp_conn.close()
        #运行服务器的update_db_in_bulk_blocker.py,目前让服务器每天晚上10点去更新
        # ssh = paramiko.SSHClient()
        # ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        # ssh.connect(hostname=host,username=user_name,password=password,key_filename='id_rsa')
        # _, stdout, stderr = ssh.exec_command('nohup python3 "/home/metron/xinyyang/blocker_status_update/update_db_in_bulk_blocker.py" &')
        # ssh.close()
        # os.path.getmtime()
        #
        # #source by source的把结果更新到数据库中
        # prompt = ''
        # for source in available_source_list:
        #
        #     print(source)
        #     # asin_blocker_info_list = all_source_df[all_source_df['source']==source][['new_blocker','new_status','asin','source']].drop_duplicates(subset=['asin']).to_numpy().tolist()
        #     asin_blocker_info_df = all_source_df[all_source_df['source']==source][['new_blocker','new_status','asin','source']].drop_duplicates(subset=['asin'])
        #     asin_blocker_info_df.fillna('',inplace=True)
        #     #把数据分成小段
        #     #分段
        #     asin_blocker_info_df.reset_index(drop=True, inplace=True)
        #     asins_by_source_num = asin_blocker_info_df.shape[0]
        #     unit_quality = 1000
        #     unit_num = math.ceil(asins_by_source_num / unit_quality)
        #     temp_df_list = []
        #     start_time = time.time()
        #     for index in range(unit_num):
        #         temp_df_list.append(asin_blocker_info_df.loc[index * unit_quality:(index + 1) * unit_quality])
        #     # print(temp_df_list)
        #     serverhandler = ServerHandler(database='notconveyable_' + source.lower())
        #     for index in range(len(temp_df_list)):
        #         asin_blocker_info_list = temp_df_list[index].to_numpy().tolist()
        #         print(get_current_progress(index, len(temp_df_list), source))
        #         serverhandler.update_blocker_values(asin_with_blocker_info_list=asin_blocker_info_list)
        #
        #     serverhandler.close_connection()
        #     end_time = time.time()
        #     print('更新'+source+'数据库用时：'+str(end_time - start_time))
        #     #分段
        #     # asin_operated_num_by_source = all_source_df[all_source_df['source']==source].drop_duplicates(subset=['asin']).shape[0]
        #     prompt += (source + '更新blocker status成功，共refresh blocker status ' + str(len(asin_blocker_info_list)) +'个;\n')
        # if prompt !='':
        #     print(prompt)
        # else:
        #     print('所有source都未成功更新 blocker status，请重试或联系xinyyang')

def get_operation_num(df,source_list):
    # print(df)
    asin_num_by_source_dic = {}
    # source_list = df['source'].unique().tolist()
    if not df.empty:
        for source in source_list:
            # asin_num_by_source_dic.append({source:df[df['source']==source].shape[0]})
            asin_num_by_source_dic[source] = df[df['source']==source].shape[0]
    else:
        for source in source_list:
            asin_num_by_source_dic[source] = 0
    # print(asin_num_by_source_dic)
    return asin_num_by_source_dic

# refresh_dimension([],[],[],r'C:\Users\xinyyang\AppData\Local\Google\Chrome\User Data',True)



def get_necessary_info_query(all_asins_df, task_name):
    '''
    根据config表中的query和每个query最多跑多少ASIN以及数据库中ASIN的数量
    整理出source list; asin_list的list; exportable_query_link list,并且return 三个list
    :return: source_list; asin_list_list; exportable_query_link_list.
    [['DE','DE'],['UK'],['US']]
    [[['1111111111','1122222222','1111222222'],['1111111111']],]
    [['http1','http2'],['http3'],['http4']]
    '''
    if task_name =='dimension':
        query_info_df = pd.read_excel(config_path,sheet_name='refresh dimension')
    elif task_name == 'blocker status':
        query_info_df = pd.read_excel(config_path,sheet_name='refresh blocker status')



    #将source转成marketplace_id
    all_asins_df.rename(columns={'source':'marketplace_id'},inplace=True)
    all_asins_df['marketplace_id'] = all_asins_df['marketplace_id'].map(lambda x:country_code_mapping_dic[x])
    final_source_list = []

    final_asin_list_list = []

    final_exportable_query_link_list = []
    # pdb.set_trace()
    #有两列，一列是marketplace_id,一列是counts，counts列记录了ASIN的数量
    asins_by_source_df = all_asins_df.groupby(['marketplace_id'])['asin'].count().reset_index(name='counts')
    # pdb.set_trace()
    source_querynumber_dic = {}
    # {4:}
    # pdb.set_trace()
    for asins_by_source_index, asins_by_source_row in asins_by_source_df.iterrows():
        source_list = []
        asin_list_list = []
        exportable_query_link_list = []
        single_query_info_df = query_info_df[
            query_info_df['source'] == code_country_mapping_dic[asins_by_source_row['marketplace_id']]]
        all_asins_single_source_df = all_asins_df[
            all_asins_df['marketplace_id'] == asins_by_source_row['marketplace_id']]
        all_asins_single_source_df.reset_index(drop=True, inplace=True)
        # pdb.set_trace()
        for query_info_index, query_info_row in single_query_info_df.iterrows():
            # pdb.set_trace()
            # total_asin_count_by_source = asins_by_source_df.loc[
            #     asins_by_source_df['marketplace_id'] == country_code_mapping_dic[
            #         query_info_row['source']]].reset_index(drop=True).loc[0, 'counts']
            total_asin_count_by_source = all_asins_single_source_df.shape[0]
            # print(type(total_asin_count_by_source))
            if total_asin_count_by_source > 0:

                leftover_num = total_asin_count_by_source - query_info_row['ASINs per query']
                # 取前query_info_row['ASINs per query']个asin，然后从df中删除这些asin
                all_asins_single_source_df.reset_index(drop=True, inplace=True)
                asins_list = all_asins_single_source_df.head(query_info_row['ASINs per query'])['asin'].tolist()
                source_list.append(query_info_row['source'])
                asin_list_list.append(asins_list)
                exportable_query_link_list.append(query_info_row['refresh query link'])
                all_asins_single_source_df = all_asins_single_source_df.loc[query_info_row['ASINs per query']:, :]
                # pdb.set_trace()
                if leftover_num <= 0:
                    # pdb.set_trace()
                    # asins_by_source_df.loc[
                    #     asins_by_source_df['marketplace_id'] == self.marketplaceid_source_mapping[
                    #         query_info_row['source']],'counts'] = 0
                    asins_by_source_df.at[asins_by_source_df.index[(
                                asins_by_source_df['marketplace_id'] == country_code_mapping_dic[
                            query_info_row['source']])], 'counts'] = 0
                    # pdb.set_trace()
                    break

                else:
                    asins_by_source_df.at[asins_by_source_df.index[(
                                asins_by_source_df['marketplace_id'] == country_code_mapping_dic[
                            query_info_row['source']])], 'counts'] = leftover_num
                    # pdb.set_trace()
        final_source_list.append(source_list)
        final_asin_list_list.append(asin_list_list)
        final_exportable_query_link_list.append(exportable_query_link_list)
        # 判断一下是否整除，若整除，对counts数取模后不加1，若不整除，取模后count数加1

    # pdb.set_trace()
    return final_source_list, final_asin_list_list, final_exportable_query_link_list

if __name__ == '__main__':
    #根据config表格中的sheet basic info中的source去进行操作，然后把对应的chrome path给到函数，让用户选择更新三个属性中的哪一个
    #然后运行程序，等待更新完成后
    source_list = config_file['source'].unique().tolist()
    proceed_flag = input('目前将要更新的国家有：'+','.join(source_list)+' . 是否对其进行更新？【Y/N】')
    if proceed_flag.lower() == 'y':
        chrome_path_list = config_file['Google user profile address'].unique().tolist()
        source_chrome_path_mapping_dic = {source_list[i]:chrome_path_list[i] for i in range(len(chrome_path_list))}
        task_flag_dic = {'refresh dimension':False,'refresh exportable date':False,'refresh blocker status':False}
        task_dic = {'refresh dimension':refresh_dimension,'refresh exportable date':refresh_exportable_date,'refresh blocker status':refresh_blocker_status}
        #source_to_refresh_list[0]中存的是要去refresh dimension的source，source_to_refresh_list[1]中存到是要去refresh exportable date的source，source_to_refresh_list[2]中存的是refresh blocker status的source
        source_to_refresh_list = []
        for k,v in task_flag_dic.items():
            temp_source_to_refresh_list = []
            action_flag = input('是否进行'+k+'【Y/N】')
            if action_flag.lower() == 'y':
                task_flag_dic[k] = True
            elif action_flag.lower() == 'n':
                source_to_refresh_list.append(temp_source_to_refresh_list)
                continue
            else:
                print('输入错误，请重新运行程序')
                sys.exit()
            #让大家输入每个source要做的task内容
            for source in source_list:
                source_action_flag = input('是否对source：'+source+'进行'+k)
                if source_action_flag.lower() == 'y':
                    temp_source_to_refresh_list.append(source)
                elif source_action_flag.lower() == 'n':
                    continue
                else:
                    print('输入错误，请重新运行程序')
                    sys.exit()
            source_to_refresh_list.append(temp_source_to_refresh_list)

        chrome_path_list = config_file['Google user profile address'].unique().tolist()
        #整理出最多三个task的functions的arguments：asin source query
        if task_flag_dic['refresh exportable date'] == True:

            pass
        if task_flag_dic['refresh dimension'] == True:
            #从数据库中抓取要查询的asin
            run_query_asins_df_list = []

            for source in source_to_refresh_list[0]:
                serverhandler = ServerHandler(database='notconveyable_'+source.lower())
                df_by_source = serverhandler.get_asins_to_check_dimension()
                run_query_asins_df_list.append(df_by_source)
            run_query_asins_df = pd.concat(run_query_asins_df_list, ignore_index=True)
            # pdb.set_trace()
            #将待查asin的df拆成source，asin，query的形式
            final_source_list_for_dimension, final_asin_list_list_for_dimension, final_exportable_query_link_list_for_dimension = get_necessary_info_query(run_query_asins_df,'dimension')
            #准备将final_source_list；final_asin_list_list；final_exportable_query_link_list；chrome_path_list放到查询的函数里
            # pdb.set_trace()
        if task_flag_dic['refresh blocker status'] == True:
            # 从数据库中抓取要查询的asin
            run_query_asins_df_list = []
            for source in source_to_refresh_list[2]:
                serverhandler = ServerHandler(database='notconveyable_' + source.lower())

                df_by_source = serverhandler.get_refresh_status_info()
                # print(df_by_source)
                run_query_asins_df_list.append(df_by_source)
            run_query_asins_df = pd.concat(run_query_asins_df_list, ignore_index=True)
            # 将待查asin的df拆成source，asin，query的形式
            # print(run_query_asins_df)
            final_source_list_for_blocker_status, final_asin_list_list_for_blocker_status, final_exportable_query_link_list_for_blocker_status = get_necessary_info_query(
                run_query_asins_df, 'blocker status')
        run_query_asins_df.to_excel('task_asins.xlsx',index=False)
        # pdb.set_trace()
        #将最多三个task放到三个进程中
        # pool_list = []
        pool = Pool(processes=3)
        # pdb.set_trace()
        #@todo 目前只能对多个source统一从头开始运行三个功能，或者统一从中间开始
        for task_k,task_v in task_flag_dic.items():
            if task_v == True:
                # pdb.set_trace()
                if task_k == 'refresh dimension':
                    result = pool.apply_async(task_dic[task_k],[final_source_list_for_dimension,final_asin_list_list_for_dimension,final_exportable_query_link_list_for_dimension,chrome_path_list[0],False])
                elif task_k == 'refresh exportable date':
                    result = pool.apply_async(task_dic[task_k],[source_to_refresh_list[1]])
                elif task_k == 'refresh blocker status':
                    result = pool.apply_async(task_dic[task_k],[final_source_list_for_blocker_status,final_asin_list_list_for_blocker_status,final_exportable_query_link_list_for_blocker_status,chrome_path_list[-1],True])
        result.get()
        pool.close()
        pool.join()

    else:
        print('未进行任何操作，程序结束')
        sys.exit()

